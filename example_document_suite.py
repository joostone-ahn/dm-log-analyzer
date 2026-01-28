#!/usr/bin/env python3
"""
Document Suite Power 예제
- Word 문서에서 텍스트 추출
- Excel 리포트 생성
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import zipfile
from xml.etree import ElementTree as ET
from datetime import datetime

def extract_text_from_docx(docx_path):
    """Word 문서에서 텍스트 추출 (OOXML 방식)"""
    try:
        with zipfile.ZipFile(docx_path, 'r') as docx:
            xml_content = docx.read('word/document.xml')
            tree = ET.XML(xml_content)
            
            # 네임스페이스 정의
            ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
            
            # 모든 텍스트 추출
            paragraphs = []
            for paragraph in tree.findall('.//w:p', ns):
                texts = [node.text for node in paragraph.findall('.//w:t', ns) if node.text]
                if texts:
                    paragraphs.append(''.join(texts))
            
            return paragraphs
    except Exception as e:
        return [f"Error: {str(e)}"]

def create_excel_report(data, output_path):
    """Excel 리포트 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Document Analysis"
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # 헤더 작성
    ws['A1'] = "Line #"
    ws['B1'] = "Content"
    ws['C1'] = "Length"
    
    for cell in ['A1', 'B1', 'C1']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
    
    # 데이터 작성
    for idx, line in enumerate(data, start=2):
        ws[f'A{idx}'] = idx - 1
        ws[f'B{idx}'] = line[:100] + "..." if len(line) > 100 else line
        ws[f'C{idx}'] = len(line)
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 10
    
    # 통계 추가
    stats_row = len(data) + 3
    ws[f'A{stats_row}'] = "Total Lines:"
    ws[f'B{stats_row}'] = len(data)
    ws[f'A{stats_row}'].font = Font(bold=True)
    
    ws[f'A{stats_row+1}'] = "Generated:"
    ws[f'B{stats_row+1}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    wb.save(output_path)
    print(f"✅ Excel report created: {output_path}")

if __name__ == "__main__":
    # 예제: specs 폴더의 Word 문서 분석
    docx_file = "specs/24501-ic0.docx"
    output_file = "document_analysis.xlsx"
    
    print(f"📄 Extracting text from: {docx_file}")
    paragraphs = extract_text_from_docx(docx_file)
    
    print(f"📊 Found {len(paragraphs)} paragraphs")
    print(f"📝 Creating Excel report...")
    
    create_excel_report(paragraphs, output_file)
    
    # 처음 3개 단락 미리보기
    print("\n📖 Preview (first 3 paragraphs):")
    for i, para in enumerate(paragraphs[:3], 1):
        preview = para[:100] + "..." if len(para) > 100 else para
        print(f"  {i}. {preview}")
